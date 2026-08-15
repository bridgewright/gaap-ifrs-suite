"""Render the three deliverables (Excel) + a machine-readable JSON."""
import json
import os
from dataclasses import asdict
import openpyxl
from .reconcile import build_reconciliation
from .difference_report import build_markdown
from .basis_grounding import load_corpus_for_grounding


def _write_financials(result, path):
    wb = openpyxl.Workbook()
    first = True
    for name, stmt in (("IFRS_BS", result.ifrs_bs), ("IFRS_PL", result.ifrs_pl)):
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = name
        ws.append(["구분", "IFRS 계정", "금액"])
        for section, accs in stmt.items():
            for acc, amt in accs.items():
                ws.append([section, acc, round(amt)])
    wb.save(path)


def _write_reconciliation(result, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    ws.append(["종류", "항목/소스", "IFRS계정/분개", "자본영향/금액", "flagged",
               "기준서(출처)", "confidence", "비고"])
    for row in build_reconciliation(result.trial_balance, result.mapped, result.adjustments):
        if row["kind"] == "reclass":
            ws.append(["재분류", row["source"], row["ifrs_account"], round(row["amount"]),
                       "", row["standard"], "high", row.get("flag", "")])
        elif row["kind"] == "adjustment":
            ws.append(["조정", row["item"], row.get("entries", ""), round(row["equity_effect"]),
                       "Y" if row.get("flagged") else "", row["standard"],
                       row.get("confidence", ""), row.get("note", "")])
        else:  # bridge
            summary = f"소스자본 {row['source_equity']:,.0f} + 조정 {row['adjustments']:,.0f}"
            ws.append(["브릿지", row["item"], summary, round(row["ifrs_equity"]),
                       "", "K-IFRS 1101", "", ""])
    wb.save(path)


def _write_impact(result, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Impact"
    ws.append(["지표", "소스GAAP", "IFRS", "델타", "%"])
    for k, v in result.impact["metrics"].items():
        ws.append([k, round(v["source"]), round(v["ifrs"]), round(v["delta"]), v["pct"]])
    ws.append([])
    ws.append(["서술", result.impact["narrative"]])
    wb.save(path)


def write_all(result, outdir, corpus_dir=None):
    os.makedirs(outdir, exist_ok=True)
    corpus = load_corpus_for_grounding(corpus_dir)
    paths = {
        "financials": os.path.join(outdir, "ifrs_financials.xlsx"),
        "reconciliation": os.path.join(outdir, "reconciliation.xlsx"),
        "impact": os.path.join(outdir, "impact_analysis.xlsx"),
        "difference": os.path.join(outdir, "difference_analysis.md"),
        "json": os.path.join(outdir, "result.json"),
    }
    _write_financials(result, paths["financials"])
    _write_reconciliation(result, paths["reconciliation"])
    _write_impact(result, paths["impact"])
    with open(paths["difference"], "w", encoding="utf-8") as f:
        f.write(build_markdown(result, corpus))
    with open(paths["json"], "w", encoding="utf-8") as f:
        json.dump({
            "ifrs_bs": result.ifrs_bs,
            "ifrs_pl": result.ifrs_pl,
            "adjustments": [asdict(a) for a in result.adjustments],
            "impact": result.impact,
        }, f, ensure_ascii=False, indent=2, default=str)
    return paths
